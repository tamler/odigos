"""Data table tool -- agent-managed structured data stored in SQLite.

The agent can create named tables, add rows, query data, and summarize.
Data lives in the DB, not on disk. Export to .xlsx is on-demand only.
This is how the agent tracks budgets, expenses, logs, habits, or any
structured data the user wants to accumulate over time.
"""
from __future__ import annotations

import json
import logging
import uuid
from datetime import datetime, timezone
from typing import TYPE_CHECKING

from odigos.tools.base import BaseTool, ToolContract, ToolResult

if TYPE_CHECKING:
    from odigos.db import Database

logger = logging.getLogger(__name__)


class DataTableTool(BaseTool):
    """Store, retrieve, and manage structured data for the user."""

    name = "data_table"
    category = "create"
    description = (
        "Create and manage structured data tables for the user. Use this to track "
        "budgets, expenses, habits, reading lists, workout logs, inventory, contacts, "
        "or any structured data the user wants to accumulate over time. "
        "Data is stored persistently and can be queried, summarized, or exported. "
        "Use action=create to set up a new table with columns. "
        "Use action=add to insert rows. Use action=query to read data back. "
        "Use action=summary for quick stats. Use action=export to generate a downloadable file. "
        "Do not use for one-off text content — use create_artifact for that."
    )
    contract = ToolContract(timeout_seconds=30)
    parameters_schema = {
        "type": "object",
        "properties": {
            "action": {
                "type": "string",
                "enum": ["create", "add", "query", "summary", "list", "delete", "export"],
                "description": "Operation to perform.",
            },
            "table": {
                "type": "string",
                "description": "Table name (e.g. 'budget', 'reading_list', 'workouts').",
            },
            "description": {
                "type": "string",
                "description": "What this table tracks (for create action).",
            },
            "columns": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Column names for create action (e.g. ['date', 'store', 'amount', 'category']).",
            },
            "rows": {
                "type": "array",
                "items": {"type": "array"},
                "description": "Data rows to add. Each row is an array matching column order.",
            },
            "limit": {
                "type": "integer",
                "description": "Max rows to return for query (default 50).",
            },
            "filter_column": {
                "type": "string",
                "description": "Column name to filter on for query.",
            },
            "filter_value": {
                "type": "string",
                "description": "Value to match in filter_column.",
            },
        },
        "required": ["action"],
    }

    def __init__(self, db: Database) -> None:
        self._db = db

    async def execute(self, params: dict) -> ToolResult:
        action = params.get("action", "")
        table_name = params.get("table", "")
        params.pop("_conversation_id", None)
        params.pop("_goal_id", None)

        if action == "list":
            return await self._list_tables()
        if action == "create":
            return await self._create_table(table_name, params)
        if not table_name:
            return ToolResult(success=False, data="", error="table name required")
        if action == "add":
            return await self._add_rows(table_name, params)
        if action == "query":
            return await self._query(table_name, params)
        if action == "summary":
            return await self._summary(table_name)
        if action == "delete":
            return await self._delete_table(table_name)
        if action == "export":
            return await self._export(table_name, params)
        return ToolResult(success=False, data="", error=f"Unknown action: {action}")

    async def _list_tables(self) -> ToolResult:
        rows = await self._db.fetch_all(
            "SELECT name, description, columns, "
            "(SELECT COUNT(*) FROM data_rows WHERE table_id = data_tables.id) as row_count "
            "FROM data_tables ORDER BY name",
        )
        if not rows:
            return ToolResult(success=True, data="No data tables exist yet. Use action=create to start one.")
        lines = ["## Your Data Tables\n"]
        for r in rows:
            cols = json.loads(r["columns"])
            desc = f" -- {r['description']}" if r["description"] else ""
            lines.append(f"**{r['name']}**{desc} ({r['row_count']} rows, columns: {', '.join(cols)})")
        return ToolResult(success=True, data="\n".join(lines))

    async def _create_table(self, name: str, params: dict) -> ToolResult:
        if not name:
            return ToolResult(success=False, data="", error="table name required")
        columns = params.get("columns", [])
        if not columns:
            return ToolResult(success=False, data="", error="columns required for create")
        description = params.get("description", "")
        now = datetime.now(timezone.utc).isoformat()
        table_id = uuid.uuid4().hex[:16]
        try:
            await self._db.execute(
                "INSERT INTO data_tables (id, name, description, columns, created_at, updated_at) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (table_id, name, description, json.dumps(columns), now, now),
            )
        except Exception as e:
            if "UNIQUE" in str(e):
                return ToolResult(success=False, data="", error=f"Table '{name}' already exists")
            raise
        return ToolResult(
            success=True,
            data=f"Created table '{name}' with columns: {', '.join(columns)}",
        )

    async def _add_rows(self, table_name: str, params: dict) -> ToolResult:
        table = await self._db.fetch_one(
            "SELECT id, columns FROM data_tables WHERE name = ?", (table_name,),
        )
        if not table:
            return ToolResult(success=False, data="", error=f"Table '{table_name}' not found")
        rows = params.get("rows", [])
        if not rows:
            return ToolResult(success=False, data="", error="rows required for add")
        now = datetime.now(timezone.utc).isoformat()
        for row in rows:
            row_id = uuid.uuid4().hex[:16]
            await self._db.execute(
                "INSERT INTO data_rows (id, table_id, values_json, created_at) VALUES (?, ?, ?, ?)",
                (row_id, table["id"], json.dumps(row), now),
            )
        await self._db.execute(
            "UPDATE data_tables SET updated_at = ? WHERE id = ?", (now, table["id"]),
        )
        return ToolResult(success=True, data=f"Added {len(rows)} row(s) to '{table_name}'.")

    async def _query(self, table_name: str, params: dict) -> ToolResult:
        table = await self._db.fetch_one(
            "SELECT id, columns FROM data_tables WHERE name = ?", (table_name,),
        )
        if not table:
            return ToolResult(success=False, data="", error=f"Table '{table_name}' not found")
        columns = json.loads(table["columns"])
        limit = params.get("limit", 50)
        filter_col = params.get("filter_column")
        filter_val = params.get("filter_value")

        rows = await self._db.fetch_all(
            "SELECT values_json FROM data_rows WHERE table_id = ? ORDER BY created_at DESC LIMIT ?",
            (table["id"], limit),
        )

        parsed = [json.loads(r["values_json"]) for r in rows]

        # Apply filter if specified
        if filter_col and filter_val and filter_col in columns:
            col_idx = columns.index(filter_col)
            parsed = [r for r in parsed if len(r) > col_idx and str(r[col_idx]).lower() == filter_val.lower()]

        if not parsed:
            return ToolResult(success=True, data=f"No data in '{table_name}'" + (f" matching {filter_col}={filter_val}" if filter_col else "") + ".")

        # Format as markdown table
        lines = [f"## {table_name} ({len(parsed)} rows)\n"]
        lines.append("| " + " | ".join(columns) + " |")
        lines.append("| " + " | ".join(["---"] * len(columns)) + " |")
        for row in parsed:
            padded = row + [""] * (len(columns) - len(row))
            lines.append("| " + " | ".join(str(c) for c in padded[:len(columns)]) + " |")
        return ToolResult(success=True, data="\n".join(lines))

    async def _summary(self, table_name: str) -> ToolResult:
        table = await self._db.fetch_one(
            "SELECT id, columns, description FROM data_tables WHERE name = ?", (table_name,),
        )
        if not table:
            return ToolResult(success=False, data="", error=f"Table '{table_name}' not found")
        columns = json.loads(table["columns"])
        count = await self._db.fetch_one(
            "SELECT COUNT(*) as cnt FROM data_rows WHERE table_id = ?", (table["id"],),
        )
        total_rows = count["cnt"] if count else 0

        # Try to compute numeric summaries
        all_rows = await self._db.fetch_all(
            "SELECT values_json FROM data_rows WHERE table_id = ?", (table["id"],),
        )
        parsed = [json.loads(r["values_json"]) for r in all_rows]

        lines = [f"## {table_name} Summary"]
        if table["description"]:
            lines.append(f"*{table['description']}*")
        lines.append(f"Total rows: {total_rows}")
        lines.append(f"Columns: {', '.join(columns)}\n")

        # Find numeric columns and compute stats
        for col_idx, col_name in enumerate(columns):
            nums = []
            for row in parsed:
                if col_idx < len(row):
                    try:
                        nums.append(float(row[col_idx]))
                    except (ValueError, TypeError):
                        pass
            if nums:
                lines.append(f"**{col_name}**: sum={sum(nums):.2f}, avg={sum(nums)/len(nums):.2f}, min={min(nums):.2f}, max={max(nums):.2f}")

        return ToolResult(success=True, data="\n".join(lines))

    async def _delete_table(self, table_name: str) -> ToolResult:
        table = await self._db.fetch_one(
            "SELECT id FROM data_tables WHERE name = ?", (table_name,),
        )
        if not table:
            return ToolResult(success=False, data="", error=f"Table '{table_name}' not found")
        await self._db.execute("DELETE FROM data_rows WHERE table_id = ?", (table["id"],))
        await self._db.execute("DELETE FROM data_tables WHERE id = ?", (table["id"],))
        return ToolResult(success=True, data=f"Deleted table '{table_name}' and all its data.")

    async def _export(self, table_name: str, params: dict) -> ToolResult:
        """Export table to .xlsx file on demand."""
        table = await self._db.fetch_one(
            "SELECT id, columns FROM data_tables WHERE name = ?", (table_name,),
        )
        if not table:
            return ToolResult(success=False, data="", error=f"Table '{table_name}' not found")

        try:
            import openpyxl
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            return ToolResult(success=False, data="", error="openpyxl not installed for export")

        columns = json.loads(table["columns"])
        rows = await self._db.fetch_all(
            "SELECT values_json FROM data_rows WHERE table_id = ? ORDER BY created_at",
            (table["id"],),
        )

        from odigos.storage import FILES_DIR
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name

        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row_idx, row in enumerate(rows, 2):
            values = json.loads(row["values_json"])
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = max(12, len(str(columns[col - 1])) + 4)

        import secrets
        filename = f"{table_name}_{secrets.token_hex(4)}.xlsx"
        filepath = str(FILES_DIR / filename)
        wb.save(filepath)

        import os
        file_size = os.path.getsize(filepath)
        now = datetime.now(timezone.utc).isoformat()
        artifact_id = secrets.token_hex(8)
        await self._db.execute(
            "INSERT OR REPLACE INTO artifacts "
            "(id, filename, content_type, file_size, file_path, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (artifact_id, filename,
             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
             file_size, filepath, now),
        )

        return ToolResult(
            success=True,
            data=f"Exported '{table_name}' to {filename} ({len(rows)} rows). Available for download.",
            side_effect={
                "artifact": {
                    "id": artifact_id,
                    "filename": filename,
                    "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "file_size": file_size,
                    "download_url": f"/api/artifacts/{artifact_id}/download",
                },
            },
        )
